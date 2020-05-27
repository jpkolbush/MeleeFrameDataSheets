import os, json, csv, math
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

JSON_PATH = 'framedata-json/'
MOVE_DICT = {
    'jab1': "Jab 1",
    'jab2': "Jab 2",
    'jab3': "Jab 3",
    'rapidjabs_start': "Rapid Jabs Startup",
    'rapidjabs_loop': "Rapid Jabs Loop",
    'rapidjabs_end': "Rapid Jabs End",
    'dashattack': "Dash Attack",
    'ftilt_m': "F-tilt",
    'utilt': "U-tilt",
    'dtilt': "D-tilt",
    'fsmash_m': "F-Smash",
    'usmash': "U-Smash",
    'dsmash': "D-Smash",
    'nair': "Nair",
    'fair': "Fair",
    'bair': "Bair",
    'uair': "Uair",
    'dair': "Dair",
    'grab': "Grab",
    'dashgrab': "Dash Grab",
    'pummel': "Pummel",
    'fthrow': "F-Throw",
    'bthrow': "B-Throw",
    'uthrow': "U-Throw",
    'dthrow': "D-Throw"
}
MAIN_SHEET = [
    ["Github page"],
    ["https://github.com/jpkolbush/MeleeFrameDataSheets"],
    [""],
    ["All data taken from here"],
    ["https://github.com/pfirsich/meleeFrameDataExtractor"],
    [""],
    ["And here"],
    ["https://smashboards.com/threads/stratocasters-hitbox-system-new-download-link.283973/"],
    [""],
    ["Hitboxes & Frame advantage"],
    ["Each range under 'Active Hits' corresponds with a range of frames where there are one or more hitboxes active"],
    ["Using Bowsers upsmash for an example, it has active hitboxes on frames 16-21 and an active hitbox on frame 30"],
    ["Under 'Damage' you can see that for bowsers up smash it has '(20, 17) | 12' listed "],
    ["This means that during the 16-21 frame window he has two hitboxes: one that does 20% damage, and one that does 17% damage. And on frame 30 he has one hitbox that does 12% damage"],
    ["Each of these hitboxes corresponds with a value in the 'Frame Advntage' column'"],
    ["Continuing with this example, '(-27,-28) | -16' is listed for the frame advantage on this move"],
    ["This means that if Bowser's frame 30 upsmash hitbox hits sheild, it is -16. Meaning that the defender has 16 frames to punish"],
    ["These frame advanatage values should be looked at with some descresion, many indicate impossible scenarios or don't show the full picture"],
    ["For example, if Bowser hits a sheild with the 20% damage hitbox on frame 16 then the defender has 27 frames to punish, but this is assuming their punish avoids the second hit"],
    ["Also, for all moves that are active on multiple frames, it is assumed that it is hitting the shield on the first active frame (frame 16 in this case)"],
    ["For all moves who's animation can be interupted early (IASA is listed), this value is used instead of 'Total Frames'"],
    [""],
    ["Arials gets even more tricky to calculate"],
    ["Lets take a look at falcon's knee. It has a frame advantage listed of '+1 | -5'. This means that the strong hit of knee is +1 on sheild"],
    ["This is assuming: 1) The falcon is L-Cancelling, and 2) The falcon is hitting the sheild on the frame right before he lands"],
    ["In general, the higher up on sheild an arial hits, the less safe it is (autocancel windows are sometimes the exception to this)"],
    ["But what about weak knee being -5?"],
    ["This is again assuming that of weak knee is hitting the sheild on the frame just before landing"],
    ["But this exact scenario is almost impossible to actually happen in match, since strong knee likely would have hit sheild beore this"],
    ["For more information on calculating frame advantage, visit this page: https://smashboards.com/threads/frame-advantage-on-block.309694/"],
    [""],
    [""],
    ["Specials, dodges, & other data completed for"],
    ["Fox"],
    ["Falco"],
    ["Math"],
    ["Jigglypuff"],
    [""],
    [""],
    ["To-Do & Known issuse:"],
    ["Finish special data"],
    ["Add Invincability frames on attacks"],
    ["Fix throw damage"]
    
]

SPECIALS = load_workbook(filename="Specials.xlsx")

def generate_row(move_name, move_json_data, character_name):
    row = []
    row.append(MOVE_DICT[move_name])
    hits = []
    hit_frames_list = move_json_data.get("hitFrames")
    for hit_frame in hit_frames_list:
        if hit_frame["start"] != hit_frame["end"]:
            hits.append("{}-{}".format(hit_frame["start"], hit_frame["end"]))
        else:
             hits.append(str(hit_frame["start"]))
    row.append(" | ".join(hits))
    row.append(move_json_data.get("totalFrames"))
    row.append(move_json_data.get("iasa"))
    row.append(move_json_data.get("landingLag"))
    l_cancelled = move_json_data.get("lcancelledLandingLag")
    if (character_name == "Game & Watch" and (move_name == "uair" or move_name == "bair" or move_name == "nair")):
        l_cancelled = move_json_data.get("landingLag")
    row.append(l_cancelled)
    notes = ""
    if (move_json_data.get("autoCancelBefore") and move_json_data.get("autoCancelAfter")):
        notes +=  "Autocancel <{} {}>".format(move_json_data.get("autoCancelBefore"), move_json_data.get("autoCancelAfter"))
    elif (move_json_data.get("autoCancelBefore")):
        notes += "Autocancel <{} ".format(move_json_data.get("autoCancelBefore"))
    elif (move_json_data.get("autoCancelAfter")):
        notes += "Autocancel {}>".format(move_json_data.get("autoCancelAfter"))
    
    if (move_json_data.get("chargeFrame")):
        notes += "Charge Frame: {}".format(move_json_data.get("chargeFrame"))


    row.append(notes)

    damage = []
    sheild_stun = []
    frame_advantage = []
    throw = move_json_data.get("throw")
    if throw:
        damage.append(str(throw["damage"]))
    elif "grab" not in move_name:
        hit_frames_list = move_json_data.get("hitFrames")
        hit_box_list = move_json_data.get("hitboxes")
        for hit_frame in hit_frames_list:
            hitboxes_list = hit_frame.get("hitboxes")
            hit_frame_damage = []
            hit_frame_sheild_stun = []
            hit_frame_frame_advantage = []
            for hit_box_index in hitboxes_list:
                d = hit_box_list[hit_box_index]["damage"]
                ss = math.floor((4.45 + d) / 2.235)
                #if its an airial assume its at base of sheild
                if (l_cancelled):
                    fa = ss - l_cancelled
                else:
                    move_duration = move_json_data.get("iasa") if move_json_data.get("iasa") else move_json_data.get("totalFrames")
                    fa = ss - (move_duration - hit_frame.get("start"))
                hit_frame_damage.append(str(d))
                hit_frame_sheild_stun.append(str(ss))
                hit_frame_frame_advantage.append(str(fa))
            #after populating each list by iterating through the hitboxes, join them togetehr
            if len(hit_frame_damage) == 1:
                damage.append(hit_frame_damage[0])
                sheild_stun.append(hit_frame_sheild_stun[0])
                frame_advantage.append(hit_frame_frame_advantage[0])
            elif len(hit_frame_damage) > 1:
                damage.append("(" + ",".join(hit_frame_damage)+ ")")
                sheild_stun.append("(" + ",".join(hit_frame_sheild_stun)+ ")")
                frame_advantage.append("(" + ",".join(hit_frame_frame_advantage)+ ")")

            
    row.append(" | ".join(damage))
    row.append(" | ".join(sheild_stun))
    row.append(" | ".join(frame_advantage))

    return row

def generate_sheet(character_json, workbook):
    character_name = character_json.split(".")[0]
    if ("Nana" in character_name):
        return workbook
    print(character_name)
    with open(JSON_PATH + character_json) as f:
        json_data = json.load(f)
    
    rows = []
    for key in json_data.keys():
        if (key != "ftilt_h" and key != "ftilt_mh" and key != "ftilt_ml" and key != "ftilt_l" 
            and  key != "fsmash_h" and key != "fsmash_mh"  and key != "fsmash_ml"  and key != "fsmash_l"
            and json_data[key] != None and key[0] != '0' and 'special' not in key):
            rows.append(generate_row(key, json_data[key], character_name))

    sheet = workbook.create_sheet(character_name)
    header = ['Move', "Active Hits", "Total Frames", "IASA", "Landing Lag", "L-Cancelled", "Notes", "Base Damage", "Sheild Stun", "Frame Advantage"]
    sheet.append(header)
    for row in rows:
        sheet.append(row)
    sheet.append([])

    if character_name not in SPECIALS.sheetnames:
        return workbook
    
    #Copy over the special moves data
    specialSheet = SPECIALS[character_name]
    for i in range(2, len(list(specialSheet.rows)) + 1):
        row = specialSheet[i]
        row_copy = []
        for elem in row:
            row_copy.append(elem.value)
        sheet.append(row_copy)
    

    return workbook
def main():
    
    wb = Workbook()
    sheet = wb.active
    for line in MAIN_SHEET:
        sheet.append(line)
    json_files = os.listdir(JSON_PATH)
    json_files.sort()
    for filename in json_files:
        wb = generate_sheet(filename, wb, )
    wb.save("FrameData.xlsx")
    


if __name__ == '__main__':
    main()